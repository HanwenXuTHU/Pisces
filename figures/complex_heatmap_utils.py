import os
import pandas as pd
import numpy as np
import PyComplexHeatmap
from PyComplexHeatmap import *

modality_dict = {'SMILES':'SMILES',
                 'Graph':'Graph',
                 'Text':'Text',
                 'Drug_target':'Target',
                 'Side_effect':'Side effect',
                 'Drug_Sensitivity_NCI60':'Sensitivity',
                 '3D':'3D',
                 'Drug_Ontology':'Drug ontology'}

def load_eval_modality(source_data_path):
    import openpyxl
    wb = openpyxl.load_workbook(source_data_path, data_only=True)
    sheet = wb['Eval single modal']
    eval_modality = []
    pred_corr = np.zeros((8, 8))
    for i in range(2, 10):
        eval_modality.append(sheet.cell(row=i, column=1).value)
        for j in range(2, 10):
            if i == j:
                pred_corr[i-2, j-2] = 1
            pred_corr[i-2, j-2] = sheet.cell(row=i, column=j).value
    eval_modality = [modality_dict[mod] for mod in eval_modality]
    auprc_list, ratio_list, mod_list = [], [], []
    for i in range(11, 19):
        auprc_list.append(sheet.cell(row=i, column=3).value)
        ratio_list.append('{:.1f}%'.format(100 - 100*float(sheet.cell(row=i, column=6).value)))
        mod_list.append(modality_dict[sheet.cell(row=i, column=1).value])
    pred_corr_df = pd.DataFrame(pred_corr, index=eval_modality, columns=eval_modality)
    pred_corr_df = pred_corr_df.reset_index().melt(id_vars='index', var_name='modality1', value_name='correlations')
    auprc_ratio_df = pd.DataFrame({'auprc':auprc_list, 'ratio':ratio_list}, index=mod_list)
    return eval_modality, pred_corr_df, auprc_ratio_df

def single_modality_complex_heatmap(eval_modal, pred_corr, auprc_ratio_df, fig_suffix='.png'):
    import matplotlib.pyplot as plt
    import matplotlib as mpl
    plt.rcParams['figure.dpi'] = 100
    plt.rcParams['savefig.dpi']=300
    is_black_background = True
    if is_black_background:
        plt.style.use('dark_background')
        mpl.rcParams.update({"ytick.color" : "w",
                            'text.color' : "w",
                            "xtick.color" : "w",
                            "axes.labelcolor" : "w",
                            "axes.edgecolor" : "w"})
    fig = plt.figure(figsize=(7,5))
    tr_cmap = plt.cm.tab20
    tr_cmap = [tr_cmap(i) for i in range(20)]
    cmap = 'Dark2'
    col_ha = HeatmapAnnotation(#label=anno_label(df_col.ColGroup, merge=True,rotation=45),
                           label=anno_label(auprc_ratio_df.ratio, c='w', merge=False,rotation=15),
                           Ratio_of_missingness=anno_simple(auprc_ratio_df.ratio,legend=False,height=2,color='w',
                                              add_text=False,text_kws={'color':'w','fontsize':10}),
                           AUPRC=anno_barplot(auprc_ratio_df.auprc,cmap=['#35978f']*10,linewidth=0.1,height=15,legend=False,
                                              width=0.3),
                           verbose=0,label_side='right',label_kws={'horizontalalignment':'left'})
    cm = DotClustermapPlotter(pred_corr,x='index',y='modality1',value='correlations',
            show_rownames=True, show_colnames=True,
            c='correlations', cmap='Blues',
            row_dendrogram_size=15,
            row_dendrogram=True,
            top_annotation=col_ha,
            vmin=0.5, vmax=1,
            tree_kws={'colors': ['w']*20, 'linewidth': 1.5},
            #dot_legend_kws={'labelspacing':3, 'handlelength':3, 'frameon':True, 'facecolor':'#35978f', 'labelcolor':'#c7eae5'},
            edgecolors='#c7eae5',
            legend_hpad=5)
    col_ha.axes[2][0].set_ylim(0.3, 0.45)
    this_file_dir = os.path.dirname(os.path.abspath(__file__))
    plt.savefig(os.path.join(this_file_dir, f'save_figs/eval_single_modal{fig_suffix}'))